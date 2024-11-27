import gspread
from oauth2client.service_account import ServiceAccountCredentials
import math
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import utils
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# Function to fetch student data from Google Sheet
def fetch_student_data(sheet_url):
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('D:\\sgsits\\5 Sem\\SOD\\updating google sheet\\ancient-episode-436506-i6-13ce67f2ebc3.json', scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_url(sheet_url).sheet1
        data = sheet.get_all_records(head=1)
        
        students = []
        for row in data:
            students.append({
                "name": row["Name"],
                "enrollment": row["Enrollment Number"],
                "cgpa": row["CGPA"],
                "domain": row["Preferred Domain"],
                "skills": row["Skills"],
                "preferred_group": row.get("Preferred Group", "").split(",")
            })
        return students
    except Exception as e:
        print(f"Error fetching student data: {e}")
        return []

# Function to create balanced groups based on CGPA
def create_groups(students, group_size):
    students_sorted = sorted(students, key=lambda x: x["cgpa"], reverse=True)
    num_groups = math.ceil(len(students) / group_size)
    groups = [[] for _ in range(num_groups)]
    group_cgpas = [0] * num_groups

    for i, student in enumerate(students_sorted):
        group_index = i % num_groups
        groups[group_index].append(student)
        group_cgpas[group_index] += student["cgpa"]

    avg_cgpas = [round(group_cgpas[i] / len(groups[i]), 2) for i in range(num_groups)]
    return groups, avg_cgpas

# Function to create batches from groups
def create_batches(groups, batch_size):
    num_batches = math.ceil(len(groups) / batch_size)
    batches = [groups[i:i + batch_size] for i in range(0, len(groups), batch_size)]
    return batches

# Function to save the output to an Excel file with improved formatting
def save_to_excel(groups, avg_cgpas, batches, filename='groups_and_batches_attractive.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Groups and Batches"

    # Title
    main_title = "Groups and Batches Report"
    ws.append([main_title])
    title_cell = ws["A1"]
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")  # White font
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="0047AB", end_color="0047AB", fill_type="solid")  # Blue background
    ws.merge_cells('A1:E1')  # Merge title across multiple columns

    # Adding borders to the title cell
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    title_cell.border = thin_border

    ws.append([])  # Blank line

    # Writing Groups and their Average CGPAs
    ws.append(["Groups and their Average CGPAs:"])
    header_cell = ws["A3"]
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")  # Light Blue
    header_cell.border = thin_border

    row_index = 4
    for i, (group, cgpa) in enumerate(zip(groups, avg_cgpas)):
        ws.append([f"Group {i + 1}:"])
        group_title_cell = ws[f"A{row_index}"]
        group_title_cell.font = Font(bold=True, color="000000")
        group_title_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Sky Blue
        group_title_cell.border = thin_border
        row_index += 1

        for student in group:
            ws.append([f"{student['name']} (Enrollment: {student['enrollment']}, CGPA: {student['cgpa']})"])
            student_cell = ws[f"A{row_index}"]
            student_cell.font = Font(color="000000")
            student_cell.border = thin_border
            row_index += 1

        # Add the average CGPA after the group
        ws.append([f"Average CGPA: {cgpa}"])
        avg_cgpa_cell = ws[f"A{row_index}"]
        avg_cgpa_cell.font = Font(italic=True, color="000000")
        avg_cgpa_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Lighter blue
        avg_cgpa_cell.border = thin_border
        row_index += 2  # Add a blank line after each group

    # Writing Batches
    ws.append(["Batches:"])
    batch_header_cell = ws[f"A{row_index}"]
    batch_header_cell.font = Font(bold=True, color="FFFFFF")
    batch_header_cell.fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")  # Light Blue
    batch_header_cell.border = thin_border
    row_index += 1

    for i, batch in enumerate(batches):
        ws.append([f"Batch {i + 1}:"])
        batch_title_cell = ws[f"A{row_index}"]
        batch_title_cell.font = Font(bold=True, color="000000")
        batch_title_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Sky Blue
        batch_title_cell.border = thin_border
        row_index += 1

        for j, group in enumerate(batch):
            ws.append([f"Group {j + 1 + i * 4}:"])
            group_batch_cell = ws[f"A{row_index}"]
            group_batch_cell.font = Font(bold=True, color="000000")
            group_batch_cell.border = thin_border
            row_index += 1

            for student in group:
                ws.append([f"{student['name']} (Enrollment: {student['enrollment']}, CGPA: {student['cgpa']})"])
                student_batch_cell = ws[f"A{row_index}"]
                student_batch_cell.font = Font(color="000000")
                student_batch_cell.border = thin_border
                row_index += 1

            # Add average CGPA after each group
            ws.append([f"Average CGPA: {avg_cgpas[j + i * 4]}"])
            avg_batch_cgpa_cell = ws[f"A{row_index}"]
            avg_batch_cgpa_cell.font = Font(italic=True, color="000000")
            avg_batch_cgpa_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Lighter blue
            avg_batch_cgpa_cell.border = thin_border
            row_index += 2  # Add a blank line between groups in batches

    # Adjust column width for better readability, skipping merged cells
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if not isinstance(cell, openpyxl.cell.cell.MergedCell)]  # Skip merged cells
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save Excel file
    wb.save(filename)
    print(f"Excel file '{filename}' has been created successfully!")

# Function to save the output to a PDF file with enhanced styling
def save_to_pdf(groups, avg_cgpas, batches, filename='groups_and_batches_report.pdf'):
    pdf = SimpleDocTemplate(filename, pagesize=A4)
    elements = []
    
    # Title styling
    title_style = ParagraphStyle(
        name="Title",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        textColor=HexColor("#0047AB"),  # Deep Blue color
    )
    
    # Add title
    title = Paragraph("Groups and Batches Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5 * inch))  # Add some space below the title

    # Header styles for sections
    section_header_style = ParagraphStyle(
        name="SectionHeader",
        fontSize=14,
        leading=16,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        backColor=HexColor("#4682B4"),  # Light Blue background
        alignment=TA_CENTER,
        borderPadding=(4, 2, 4, 2),  # Add padding inside the header cells
    )

    # Paragraph style for student details
    student_style = ParagraphStyle(
        name="StudentDetails",
        fontSize=12,
        leading=14,
        fontName="Helvetica",
    )

    # Style for average CGPA
    avg_cgpa_style = ParagraphStyle(
        name="AverageCGPA",
        fontSize=12,
        leading=14,
        fontName="Helvetica-Oblique",
        textColor=colors.darkblue,
    )

    # Table styling for groups
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor("#ADD8E6")),  # Light Blue background for header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # White text color for header
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor("#F0F8FF")),  # Alice Blue for body
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),  # Add grey borders for all cells
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Outer border for the table
    ])

    # Add the "Groups" section
    elements.append(Paragraph("Groups and their Average CGPAs", section_header_style))
    elements.append(Spacer(1, 0.2 * inch))

    for i, (group, avg_cgpa) in enumerate(zip(groups, avg_cgpas)):
        # Group Heading Block
        group_heading_style = ParagraphStyle(
            name="GroupHeading",
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            backColor=HexColor("#B0C4DE"),  # Light Steel Blue
            borderPadding=(4, 2, 4, 2),  # Padding inside the header block
            textColor=colors.black,  # Black text color for readability
        )
        elements.append(Paragraph(f"Group {i + 1}", group_heading_style))
        elements.append(Spacer(1, 0.1 * inch))

        # Create table for each group
        table_data = [["Student Name", "Enrollment Number", "CGPA"]]
        for student in group:
            table_data.append([student['name'], student['enrollment'], student['cgpa']])

        group_table = Table(table_data, colWidths=[2 * inch, 2 * inch, 1.5 * inch])
        group_table.setStyle(table_style)
        elements.append(group_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Add average CGPA with a different styling
        elements.append(Paragraph(f"Average CGPA: {avg_cgpa}", avg_cgpa_style))
        elements.append(Spacer(1, 0.5 * inch))  # Add space after each group

    # Add a page break between sections
    elements.append(PageBreak())

    # Add the "Batches" section
    elements.append(Paragraph("Batches", section_header_style))
    elements.append(Spacer(1, 0.2 * inch))

    for i, batch in enumerate(batches):
        elements.append(Paragraph(f"Batch {i + 1}", section_header_style))
        elements.append(Spacer(1, 0.1 * inch))

        for j, group in enumerate(batch):
            elements.append(Paragraph(f"Group {j + 1 + i * 4}", student_style))
            elements.append(Spacer(1, 0.1 * inch))

            # Create table for each group in the batch
            table_data = [["Student Name", "Enrollment Number", "CGPA"]]
            for student in group:
                table_data.append([student['name'], student['enrollment'], student['cgpa']])

            batch_table = Table(table_data, colWidths=[2 * inch, 2 * inch, 1.5 * inch])
            batch_table.setStyle(table_style)
            elements.append(batch_table)
            elements.append(Spacer(1, 0.2 * inch))

            # Add average CGPA for the group
            elements.append(Paragraph(f"Average CGPA: {avg_cgpas[j + i * 4]}", avg_cgpa_style))
            elements.append(Spacer(1, 0.5 * inch))  # Add space after each group in the batch

    # Build the PDF
    pdf.build(elements)
    print(f"PDF file '{filename}' has been created successfully!")

# Main logic to run the program
sheet_url = 'https://docs.google.com/spreadsheets/d/1eVDkpxd5hfp-xDN_AzbKsfx9v5tBuj-6ssERLydcWAw/edit?resourcekey=&gid=86436192'
students = fetch_student_data(sheet_url)

# Create groups of 5 students
groups, avg_cgpas = create_groups(students, group_size=5)

# Create batches of 4 groups
batches = create_batches(groups, batch_size=4)

# Save both PDF and Excel
save_to_pdf(groups, avg_cgpas, batches)
save_to_excel(groups, avg_cgpas, batches)
